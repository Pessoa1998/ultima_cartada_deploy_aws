from django.shortcuts import render, get_object_or_404, redirect
from django.views import generic
from .models import Todo
from django.http import HttpResponseRedirect
import openpyxl

import csv

from django.http import HttpResponse
from django.shortcuts import render

class IndexView(generic.ListView):
    template_name = 'todos/index.html'
    context_object_name = 'todo_list'

    def get_queryset(self):
        """Return all the latest todos."""
        return Todo.objects.order_by('-created_at')

def add(request):
    title = request.POST['title']
    Todo.objects.create(title=title)

    return redirect('todos:index')

def delete(request, todo_id):
    todo = get_object_or_404(Todo, pk=todo_id)
    todo.delete()

    return redirect('todos:index')

def update(request, todo_id):
    todo = get_object_or_404(Todo, pk=todo_id)
    isCompleted = request.POST.get('isCompleted', False)
    if isCompleted == 'on':
        isCompleted = True
    
    todo.isCompleted = isCompleted

    todo.save()
    return redirect('todos:index')

# def export_todos_to_excel(request):
#     # Configurar o response para download
#     response = HttpResponse(content_type='text/csv')
#     response['Content-Disposition'] = 'attachment; filename="atividades.csv"'

#     writer = csv.writer(response)
#     writer.writerow(['ID', 'Atividade', 'Concluído'])  # Cabeçalhos

#     for todo in Todo.objects.all():
#         writer.writerow([todo.id, todo.title, 'Sim' if todo.isCompleted else 'Não'])

#     return response

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Todo


def export_todos_to_excel(request):
    # Cria uma nova workbook e ativa a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Atividades'

    # Adiciona o cabeçalho
    headers = ['ID', 'Atividade', 'Concluído']
    ws.append(headers)

    # Adiciona os dados
    for todo in Todo.objects.all():
        ws.append([todo.id, todo.title, 'Sim' if todo.isCompleted else 'Não'])

    # Ajusta a largura das colunas
    for col_num, column_title in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 25

    # Configura o response para download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="atividades.xlsx"'

    # Salva a workbook no response
    wb.save(response)

    return response




def todos_view(request):
    # Lógica para exibir os todos
    return render(request, 'todos/index.html')